# Importar las bibliotecas necesarias
from flask import Flask, request, jsonify
import os
from openai import OpenAI
import pandas as pd  # Para manejar datos tabulares y guardar en Excel
from openpyxl import load_workbook  # Para manejar archivos Excel existentes

# Crear una instancia de la aplicación Flask
app = Flask(__name__)

# Configuración de OpenAI
# Obtener el token de autenticación desde una variable de entorno para proteger las credenciales.
# os.getenv devuelve None si la variable no está definida.
token = os.getenv("GITHUB_TOKEN")

# Definir el endpoint de OpenAI y el modelo a usar
endpoint = "https://models.inference.ai.azure.com"
model_name = "gpt-4o"  # Modelo GPT-4 optimizado

# Verificar si el token está configurado. Si no, lanzar un error para evitar problemas de ejecución.
if not token:
    raise EnvironmentError(
        "El token de OpenAI no está configurado en las variables de entorno"
    )

# Crear un cliente de OpenAI utilizando el token y el endpoint configurados
client = OpenAI(base_url=endpoint, api_key=token)


@app.route('/')
def home():
    # Responde con un mensaje indicando que el servicio está funcionando.
    return "fillout_service is mounted"


@app.route('/process_text', methods=['POST'])
def process_text():
    """
    Endpoint para procesar texto y rellenar formularios médicos.
    """
    # Obtener los datos JSON enviados en la solicitud POST.
    data = request.get_json()

    # Validar que los datos contengan el campo 'conversation'.
    if not data or 'transcription' not in data:
        return jsonify({"error": "Datos no válidos, se esperaba el campo 'conversation'"}), 400

    # Extraer el texto del diálogo de la solicitud.
    conversation = data['transcription']

    # Construir una consulta detallada para GPT-4 con instrucciones específicas.
    query = (
        "A continuación, se te proporcionará una serie de diálogos extraídos de consultas médicas. "
        "Tu tarea es procesar esta información para completar un registro médico con los siguientes especificaciones: "
        "1. Campos a rellenar: - Nombre - Sexo (solo puede ser 'Masculino' o 'Femenino') - Edad - Motivo de la consulta "
        "- Problema actual - Antecedentes personales - Antecedentes familiares - Vacunación - Diagnóstico - Observaciones - Tratamiento (Lo más detallado posible). "
        "2. Reglas a seguir: "
        "- Si algún campo no tiene información explícita en el diálogo, completa con 'Desconocido'. "
        "- El campo 'Sexo' solo puede ser 'Masculino' o 'Femenino'. "
        "- Los campos 'Diagnóstico', 'Observaciones' y 'Tratamiento' deben completarse según tu análisis del diálogo y la información que puedas rescatar del mismo, y aunque no se mencionen textualmente. "
        "3. Formato de respuesta: Responde SOLO con los valores de los campos, separados por punto y coma (;). "
        "4. Ejemplo de entrada (Diálogo): Paciente, Juan Pérez, refiere dolor abdominal después de comer, con antecedentes de gastritis... "
        "Ejemplo de salida: Juan Pérez; Masculino; 45 años; Dolor abdominal postprandial; Gastritis; Sin antecedentes personales importantes; "
        "Madre con úlcera gástrica; Vacunación al día; Gastritis crónica; Se sugiere endoscopía y evaluación adicional; Omeprazol 20 mg cada 12 horas por 2 semanas. "
        "En caso de que no detectes un diálogo correspondiente a atención médica, simplemente responde a todos los campos con Desconocido y nada más. "
    )

    # Añadir el diálogo proporcionado por el usuario a la consulta.
    query += f"\n{conversation}"

    try:
        # Enviar la consulta al modelo GPT-4 y obtener una respuesta.
        response = client.chat.completions.create(
            messages=[
                {"role": "system", "content": "You are a knowledgeable doctor, providing accurate, and helpful medical advice based on the information given."},
                {"role": "user", "content": query}
            ],
            temperature=0.7,  # Controla la creatividad de la respuesta.
            # Muestra la probabilidad acumulativa para limitar el vocabulario generado.
            top_p=1.0,
            max_tokens=1000,  # Límite de tokens en la respuesta.
            model=model_name
        )

        # Extraer el texto de la respuesta del modelo.
        result = response.choices[0].message.content

        # Dividir la respuesta en los diferentes campos separados por punto y coma (;).
        response_data = result.split(";")

        # Guardar los datos en un archivo Excel.
        excel_path = "caracteristicas.xlsx"

        # Si el archivo Excel no existe, crear uno con las columnas necesarias.
        if not os.path.exists(excel_path):
            df = pd.DataFrame(columns=["Nombre", "Sexo", "Edad", "Motivo de la consulta", "Problema Actual", "Antecedentes Personales",
                                       "Antecedentes Familiares", "Vacunación", "Diagnóstico", "Observaciones", 'Tratamiento'])
            df.to_excel(excel_path, index=False)

        # Abrir el archivo Excel existente.
        workbook = load_workbook(excel_path)
        sheet = workbook.active

        # Determinar la siguiente fila vacía en el archivo Excel.
        next_row = sheet.max_row + 1

        # Escribir los datos en las celdas correspondientes.
        for col_num, value in enumerate(response_data, start=1):
            sheet.cell(row=next_row, column=col_num, value=value)

        # Guardar los cambios en el archivo Excel.
        workbook.save(excel_path)

        # Devolver un mensaje de éxito junto con los datos procesados.
        return jsonify({"message": "Datos procesados y guardados correctamente", "result": response_data})

    except Exception as e:
        # Capturar errores durante el procesamiento y devolver un mensaje de error.
        return jsonify({"error": str(e)}), 500


if __name__ == '__main__':
    # Ejecutar la aplicación Flask en la dirección IP y puerto especificados.
    app.run(host='26.164.147.61', port=5001, debug=True)
